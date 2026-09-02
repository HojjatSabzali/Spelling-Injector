import json
import tkinter as tk
from tkinter import messagebox, filedialog, scrolledtext, ttk
from datetime import datetime
import os
import sys
import openpyxl
import random
import pygame
from gtts import gTTS
import numpy as np
from PIL import Image, ImageTk
import pyttsx3
import shutil
import webbrowser
import threading
import ctypes
import uuid
import zipfile
import xml.etree.ElementTree as ET

# --- Determine correct path for data files ---
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

WORDS_PATH = os.path.join(SCRIPT_DIR, 'words.xlsx')
MEMORIZED_PATH = os.path.join(SCRIPT_DIR, 'memorized.xlsx')
NEW_PATH = os.path.join(SCRIPT_DIR, 'new.xlsx')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'config.json')
ICONS_DIR = os.path.join(SCRIPT_DIR, 'icons')
ICON_PATH = os.path.join(SCRIPT_DIR, 'app_icon.ico')
BACKUP_DIR = os.path.join(SCRIPT_DIR, 'backups')
PRONUNCIATIONS_DIR = os.path.join(SCRIPT_DIR, 'pronunciations')

# Create necessary directories
if not os.path.exists(ICONS_DIR): os.makedirs(ICONS_DIR)
if not os.path.exists(PRONUNCIATIONS_DIR): os.makedirs(PRONUNCIATIONS_DIR)

# --- Config Management ---
DEFAULT_CONFIG = {
    "target_correct": 5,
    "tts_mode": "auto",  # auto, offline
    "accent": "us",      # us, uk
    "hide_add_word_help": False,
    "tts_speed": "normal", # slow, normal, fast
    "theme_color": "#2A0845" # Default Dark Purple
}

def load_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r') as f:
                config = json.load(f)
                for k, v in DEFAULT_CONFIG.items():
                    if k not in config: config[k] = v
                return config
        except: pass
    return DEFAULT_CONFIG.copy()

def save_config(config):
    with open(CONFIG_PATH, 'w') as f:
        json.dump(config, f)

# --- Dynamic Theme Colors Logic ---
def get_luminance(hex_color):
    hex_color = hex_color.lstrip('#')
    if len(hex_color) == 6:
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        return (0.299 * r + 0.587 * g + 0.114 * b)
    return 0

def adjust_color(hex_color, factor):
    hex_color = hex_color.lstrip('#')
    if len(hex_color) != 6: return "#808080"
    r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    r = min(max(int(r * factor), 0), 255)
    g = min(max(int(g * factor), 0), 255)
    b = min(max(int(b * factor), 0), 255)
    return f"#{r:02x}{g:02x}{b:02x}"

# Global variables to be set dynamically
BG_COLOR = "#2A0845"
FG_COLOR = "#E0E0E0"
ENTRY_BG = "#3E1E54"
ENTRY_FG = "#FFFFFF"
BTN_BG = "#4A1942"
BTN_FG = "#FFFFFF"
ACCENT_COLOR = "#D4AF37"

def apply_theme():
    global BG_COLOR, FG_COLOR, ENTRY_BG, ENTRY_FG, BTN_BG, ACCENT_COLOR
    config = load_config()
    bg = config.get("theme_color", "#2A0845")
    BG_COLOR = bg
    
    lum = get_luminance(bg)
    if lum > 140: # Light Theme chosen
        FG_COLOR = "#000000"
        ENTRY_FG = "#000000"
        # If very bright, darken entry background slightly. If medium bright, lighten it.
        ENTRY_BG = adjust_color(bg, 1.1) if lum < 220 else adjust_color(bg, 0.9)
        BTN_BG = adjust_color(bg, 0.8) # Darken buttons slightly for contrast
        ACCENT_COLOR = "#B8860B" # Darker gold for visibility on light bg
    else: # Dark Theme chosen
        FG_COLOR = "#E0E0E0"
        ENTRY_FG = "#FFFFFF"
        ENTRY_BG = adjust_color(bg, 1.4) # Lighten slightly for entry box
        BTN_BG = adjust_color(bg, 1.5)   # Lighten slightly for buttons
        ACCENT_COLOR = "#D4AF37" # Standard Gold

apply_theme()

# Matrix of 50 colors (10x5) for the Color Picker
THEME_COLORS = [
    ["#000000", "#404040", "#808080", "#C0C0C0", "#FFFFFF", "#1A1A1A", "#2A0845", "#0F0F0F", "#333333", "#E6E6E6"],
    ["#000080", "#0000FF", "#4169E1", "#87CEEB", "#E0FFFF", "#4B0082", "#8A2BE2", "#9370DB", "#DDA0DD", "#E6E6FA"],
    ["#006400", "#008000", "#228B22", "#3CB371", "#98FB98", "#556B2F", "#8FBC8F", "#66CDAA", "#2E8B57", "#00FF7F"],
    ["#800000", "#8B0000", "#B22222", "#FF0000", "#FA8072", "#FF4500", "#FF8C00", "#FFA500", "#8B4513", "#D2691E"],
    ["#B8860B", "#DAA520", "#FFD700", "#FFFF00", "#FFFACD", "#C71585", "#FF1493", "#FF69B4", "#FFC0CB", "#FFE4E1"]
]

# --- Change Windows Titlebar Color (Windows 11 API) ---
def set_titlebar_color(win, color_hex):
    def apply_color():
        try:
            win.update_idletasks()
            # Fetch the window handle (HWND) for the desktop window manager
            hwnd = ctypes.windll.user32.GetParent(win.winfo_id())
            b = int(color_hex[5:7], 16)
            g = int(color_hex[3:5], 16)
            r = int(color_hex[1:3], 16)
            color_bgr = (b << 16) | (g << 8) | r
            # Apply the custom titlebar color using Windows 11 DWM API
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, ctypes.byref(ctypes.c_int(color_bgr)), 4)
        except Exception:
            pass
            
    # Schedule the color application after 20ms to ensure all window properties (like transient) are fully loaded
    win.after(20, apply_color)

# --- Data Management ---
def create_empty_db_files():
    if not os.path.exists(WORDS_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'words'
        ws.append(["Word", "Correct", "Total"])
        wb.create_sheet('queue')
        wb.save(WORDS_PATH)
        wb.close()
    if not os.path.exists(MEMORIZED_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Memorized'
        ws.append(["Word", "Total Shown", "Memorized Date"])
        wb.save(MEMORIZED_PATH)
        wb.close()
    if not os.path.exists(NEW_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'New'
        ws.append(["Word"])
        wb.save(NEW_PATH)
        wb.close()

def load_data_from_file(file_path, queue_sheet_name='queue'):
    create_empty_db_files()
    wb = openpyxl.load_workbook(file_path, data_only=True)
    words_sheet = wb['words'] if 'words' in wb else wb.active
    queue_sheet = wb[queue_sheet_name] if queue_sheet_name and queue_sheet_name in wb else None

    words_dict = {}
    for row in words_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        word = row[0]
        if word:
            # Remove asterisks to get the clean word for spell checking logic
            clean_word = str(word).replace('*', '').strip().lower()
            words_dict[clean_word] = {
                "original_word": str(word), # Keeps the asterisks for display (e.g. th*ie*f)
                "correct_count": int(row[1] or 0),
                "total_count": int(row[2] or 0)
            }

    queue = []
    if queue_sheet:
        for row in queue_sheet.iter_rows(max_col=1, values_only=True):
            if row[0]:
                # Always keep queue clean of asterisks
                queue.append(str(row[0]).replace('*', '').strip().lower())

    wb.close()
    return {'words': words_dict, 'queue': queue}

def save_data_to_file(data, file_path, queue_sheet_name='queue'):
    wb = openpyxl.Workbook()
    words_sheet = wb.active
    words_sheet.title = 'words'
    words_sheet.append(["Word", "Correct", "Total"])

    for clean_word, d in data['words'].items():
        words_sheet.append([d['original_word'], d['correct_count'], d['total_count']])

    if queue_sheet_name:
        queue_sheet = wb.create_sheet(queue_sheet_name)
        for w in data['queue']:
            queue_sheet.append([w])

    wb.save(file_path)
    wb.close()

def add_to_memorized(original_word, total_count):
    create_empty_db_files()
    wb = openpyxl.load_workbook(MEMORIZED_PATH)
    sheet = wb.active
    sheet.append([original_word, total_count, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(MEMORIZED_PATH)
    wb.close()

# --- Backup and Restore ---
def backup_data():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder_path = os.path.join(BACKUP_DIR, f"Backup_{timestamp}")
    os.makedirs(folder_path)
    try:
        if os.path.exists(WORDS_PATH): shutil.copy(WORDS_PATH, folder_path)
        if os.path.exists(MEMORIZED_PATH): shutil.copy(MEMORIZED_PATH, folder_path)
        if os.path.exists(CONFIG_PATH): shutil.copy(CONFIG_PATH, folder_path)
        messagebox.showinfo("Backup Success", f"Backup created successfully at:\n{folder_path}")
        return True
    except Exception as e:
        messagebox.showerror("Backup Error", f"Failed to create backup: {e}")
        return False

def restore_data():
    folder_path = filedialog.askdirectory(title="Select Backup Folder to Restore", initialdir=BACKUP_DIR)
    if not folder_path: return
    
    w_path = os.path.join(folder_path, 'words.xlsx')
    c_path = os.path.join(folder_path, 'memorized.xlsx')
    
    if not os.path.exists(w_path):
        messagebox.showerror("Restore Error", "words.xlsx not found in the selected folder.\nPlease check Help for manual restore format.")
        return

    try:
        wb = openpyxl.load_workbook(w_path, data_only=True)
        ws = wb.active
        if ws.cell(1,1).value != "Word" or ws.cell(1,2).value != "Correct" or ws.cell(1,3).value != "Total":
            messagebox.showerror("Restore Error", "Invalid column titles in words.xlsx.\nExpected: Word | Correct | Total.\nCheck Help for instructions.")
            return
        wb.close()
    except Exception as e:
        messagebox.showerror("Restore Error", f"Cannot read backup files: {e}")
        return

    try:
        if os.path.exists(w_path): shutil.copy(w_path, WORDS_PATH)
        if os.path.exists(c_path): shutil.copy(c_path, MEMORIZED_PATH)
        cfg_path = os.path.join(folder_path, 'config.json')
        if os.path.exists(cfg_path): shutil.copy(cfg_path, CONFIG_PATH)
        messagebox.showinfo("Restore Success", "Data restored successfully!")
    except Exception as e:
        messagebox.showerror("Restore Error", f"Failed during copy: {e}")

# --- Add Word ---
def prompt_add_word_help():
    config = load_config()
    if config.get("hide_add_word_help", False): return True
    
    win = tk.Toplevel()
    win.withdraw()  # Hide window during rendering
    win.title("How to add words")
    win.geometry("680x250")
    win.configure(bg=BG_COLOR)
    set_titlebar_color(win, BG_COLOR)
    try: win.iconbitmap(ICON_PATH)
    except: pass
    win.geometry(f"+{(win.winfo_screenwidth() - 680) // 2}+{(win.winfo_screenheight() - 250) // 2}")
    win.grab_set()

    tk.Label(win, text="Adding New Words from Excel", font=("Calibri", 14, "bold"), bg=BG_COLOR, fg=FG_COLOR).pack(pady=10)
    msg = "Please ensure your Excel file has the words listed in the FIRST column (Column A).\nThe first row can be a header (e.g., 'Word') and will be skipped.\nFor a detailed guide, check the Help menu."
    tk.Label(win, text=msg, font=("Calibri", 11), bg=BG_COLOR, fg=FG_COLOR, justify=tk.LEFT).pack(pady=10, padx=20)

    var = tk.IntVar()
    cb = tk.Checkbutton(win, text="Don't show this message again", variable=var, font=("Calibri", 10), bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR)
    cb.pack(pady=10)

    res = [False]
    def proceed():
        if var.get() == 1:
            config["hide_add_word_help"] = True
            save_config(config)
        res[0] = True
        win.destroy()

    RoundedButton(win, text="Select Excel File", command=proceed, bg_color="#2196F3", fg_color="white", width=180, height=40).pack(pady=10)
    win.deiconify()  # Show window after setup is complete
    win.wait_window()
    return res[0]

def extract_rich_text_mapping(file_path):
    import xml.etree.ElementTree as ET
    import zipfile
    mapping = {}
    
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            if 'xl/sharedStrings.xml' not in z.namelist():
                return mapping
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                
                for si in root.iter():
                    if si.tag.endswith('}si') or si.tag == 'si':
                        runs = [c for c in si if c.tag.endswith('}r') or c.tag == 'r']
                        if runs:
                            plain_text = ""
                            runs_list = []
                            for r in runs:
                                t_text = ""
                                has_rpr = False
                                has_b = False
                                b_val = True
                                
                                for child in r:
                                    if child.tag.endswith('}t') or child.tag == 't':
                                        if child.text: t_text = child.text
                                    elif child.tag.endswith('}rPr') or child.tag == 'rPr':
                                        has_rpr = True
                                        for prop in child:
                                            if prop.tag.endswith('}b') or prop.tag == 'b':
                                                has_b = True
                                                val = prop.get('val')
                                                # If val is explicitly set to 0 or false, it's not bold
                                                if val in ['0', 'false', 'False']:
                                                    b_val = False
                                                    
                                if t_text:
                                    # EXCEL LOGIC REVEALED:
                                    if not has_rpr:
                                        # No properties? It inherits the global cell style (None)
                                        bold_status = None 
                                    elif has_rpr and not has_b:
                                        # Has properties but no 'b' tag? Explicitly NOT bold
                                        bold_status = False
                                    else:
                                        # Has 'b' tag, use its value
                                        bold_status = b_val
                                        
                                    plain_text += t_text
                                    runs_list.append((t_text, bold_status))
                            
                            if plain_text and len(runs_list) > 0:
                                mapping[plain_text.strip().lower()] = runs_list
    except Exception as e:
        print(f"XML parsing error: {e}")
        
    return mapping

def load_words_from_xlsx():
    if not prompt_add_word_help(): return
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not file_path: return

    try:
        # 1. First, we extract the map of bolded words directly from the heart of Excel XML.
        rich_text_dict = extract_rich_text_mapping(file_path)

        # 2. We read the file in the usual way to go row by row
        input_wb = openpyxl.load_workbook(file_path, data_only=True)
        input_sheet = input_wb.active

        data = load_data_from_file(WORDS_PATH)
        words_dict = data['words']
        queue = data['queue']
        added_count = 0
        new_words = []
        
        create_empty_db_files()
        comp_wb = openpyxl.load_workbook(MEMORIZED_PATH, data_only=True)
        memorized_words = [str(r[0]).strip().lower() for r in comp_wb.active.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
        comp_wb.close()

        # Read the first column (from the second row onwards)
        for row in input_sheet.iter_rows(min_row=2, max_col=1):
            cell = row[0]
            if cell.value:
                plain_word = str(cell.value).strip()
                lookup_key = plain_word.lower()
                
                # Check global cell font style
                cell_is_bold = bool(cell.font and cell.font.b)
                runs_list = rich_text_dict.get(lookup_key)
                
                if runs_list:
                    starred_word = ""
                    for text_part, bold_status in runs_list:
                        # If bold_status is None, it inherits from cell_is_bold. Otherwise, use explicit status.
                        is_bold = bold_status if bold_status is not None else cell_is_bold
                        
                        if is_bold:
                            starred_word += f"*{text_part}*"
                        else:
                            starred_word += text_part
                            
                    # Clean up double stars caused by adjacent bold letters (e.g. *t**h* -> *th*)
                    while "**" in starred_word:
                        starred_word = starred_word.replace("**", "")
                else:
                    # Fallback if no rich text formatting was found inside the cell runs
                    if cell_is_bold:
                        starred_word = f"*{plain_word}*"
                    else:
                        starred_word = plain_word
                
                clean_word = starred_word.replace('*', '').strip().lower()

                # Register the new word in the dictionary
                if clean_word and clean_word not in words_dict and clean_word not in memorized_words:
                    is_starred = ('*' in starred_word)
                    words_dict[clean_word] = {
                        "original_word": starred_word,
                        "correct_count": 0,
                        "total_count": 0,
                        "star": is_starred 
                    }
                    new_words.append(clean_word)
                    added_count += 1

        random.shuffle(new_words)
        queue.extend(new_words)
        save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
        input_wb.close()
        messagebox.showinfo("Success", f"{added_count} new words added and saved!\nTotal words in queue: {len(words_dict)}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not read the Excel file:\n{e}")
                
# --- Audio Management ---
def play_offline(word, config):
    try:
        engine = pyttsx3.init()
        voices = engine.getProperty('voices')
        accent = config.get("accent", "us").strip().lower() 
        
        for voice in voices:
            vid = voice.id.lower()
            vname = voice.name.lower()
            if accent == 'uk' and ('uk' in vid or 'brit' in vid or 'gb' in vid or 'great britain' in vname):
                engine.setProperty('voice', voice.id)
                break
            elif accent == 'us' and ('us' in vid or 'american' in vid or 'america' in vname):
                engine.setProperty('voice', voice.id)
                break
        
        speed_setting = config.get("tts_speed", "normal")
        if speed_setting == "slow": engine.setProperty('rate', 130)
        elif speed_setting == "fast": engine.setProperty('rate', 250)
        else: engine.setProperty('rate', 180)

        # Generate random name for offline file to prevent caching by pygame
        temp_file = os.path.join(SCRIPT_DIR, f"temp_offline_{uuid.uuid4().hex}.wav")
        engine.save_to_file(word, temp_file)
        engine.runAndWait()

        sound = pygame.mixer.Sound(temp_file)
        sound.play()
        while pygame.mixer.get_busy():
            pygame.time.Clock().tick(10)
        
        try: os.remove(temp_file)
        except: pass
        return True
    except Exception as e:
        print(f"Offline TTS error: {e}")
        return False

def play_sound(word, config, force_offline=False):
    if config["tts_mode"] == "offline" or force_offline:
        return play_offline(word, config)

    try:
        accent = config.get("accent", "us").strip().lower()
        tld_val = 'us' if accent == 'us' else 'co.uk'
        speed_setting = config.get("tts_speed", "normal")
        is_slow = (speed_setting == "slow")
        speed_suffix = "slow" if is_slow else "normal"
        
        # Clean the word and create the permanent storage path
        clean_word = word.replace('*', '').strip().lower()
        permanent_file = os.path.join(PRONUNCIATIONS_DIR, f"{clean_word}_{accent}_{speed_suffix}.mp3")
        
        # 1. Check if the file has already been downloaded and saved
        if not os.path.exists(permanent_file):
            # If not, download and save it in the permanent directory
            tts = gTTS(text=clean_word, lang='en', tld=tld_val, slow=is_slow)
            tts.save(permanent_file)
            
        # 2. Play from the local file (high speed)
        sound = pygame.mixer.Sound(permanent_file)
        sound.play()
        
        while pygame.mixer.get_busy():
            pygame.time.Clock().tick(10)
            
        return True
    except Exception as e:
        print(f"Online TTS error: {e}")
        return False

def play_success_sound():
    frequency, duration, volume = 880, 0.2, 0.3
    sample_rate = 44100
    t = np.linspace(0, duration, int(sample_rate * duration), False)
    wave = np.sin(2 * np.pi * frequency * t)
    sound_array = np.int16(wave * 32767 * volume)
    stereo_sound = np.repeat(sound_array.reshape(-1, 1), 2, axis=1)
    pygame_sound = pygame.sndarray.make_sound(stereo_sound)
    pygame_sound.play()

def play_error_sound():
    frequency, duration, volume = 440, 0.3, 0.3
    sample_rate = 44100
    t = np.linspace(0, duration, int(sample_rate * duration), False)
    wave = np.sin(2 * np.pi * frequency * t)
    sound_array = np.int16(wave * 32767 * volume)
    stereo_sound = np.repeat(sound_array.reshape(-1, 1), 2, axis=1)
    pygame_sound = pygame.sndarray.make_sound(stereo_sound)
    pygame_sound.play()

# --- Custom UI Elements ---
class RoundedButton(tk.Canvas):
    def __init__(self, parent, text, command, bg_color, fg_color, width=150, height=40, radius=20, font=("Calibri", 12)):
        super().__init__(parent, width=width, height=height, bg=parent['bg'], highlightthickness=0)
        self.command = command
        self.bg_color = bg_color
        self.original_bg = bg_color
        self.original_fg = fg_color
        self.active_color = self.lighten_color(bg_color, 1.15)
        self.is_disabled = False
        
        # Disabled colors
        self.disabled_bg = "#808080" 
        self.disabled_fg = "#D3D3D3"
        
        self.rect_id = self.create_rounded_rect(0, 0, width, height, radius, fill=bg_color)
        self.text_id = self.create_text(width/2, height/2, text=text, fill=fg_color, font=font)
        
        self.config(cursor="hand2")
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        
        self.bind("<ButtonRelease-1>", self.handle_click)
        
    def handle_click(self, event=None):
        if not getattr(self, 'is_disabled', False):
            self.command()

    def disable(self):
        self.is_disabled = True
        self.itemconfig(self.rect_id, fill=self.disabled_bg)
        self.itemconfig(self.text_id, fill=self.disabled_fg)
        self.config(cursor="arrow")

    def enable(self):
        self.is_disabled = False
        self.itemconfig(self.rect_id, fill=self.bg_color)
        self.itemconfig(self.text_id, fill=self.original_fg)
        self.config(cursor="hand2")

    def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        points = [
            x1+radius, y1, x2-radius, y1, x2, y1, x2, y1+radius,
            x2, y2-radius, x2, y2, x2-radius, y2, x1+radius, y2,
            x1, y2, x1, y2-radius, x1, y1+radius, x1, y1
        ]
        return self.create_polygon(points, smooth=True, **kwargs)

    def lighten_color(self, hex_color, factor):
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = min(int(r * factor), 255)
        g = min(int(g * factor), 255)
        b = min(int(b * factor), 255)
        return f"#{r:02x}{g:02x}{b:02x}"

    def on_enter(self, event):
        if not self.is_disabled:
            self.itemconfig(self.rect_id, fill=self.active_color)

    def on_leave(self, event):
        if not self.is_disabled:
            self.itemconfig(self.rect_id, fill=self.bg_color)

    def set_text(self, text):
        self.itemconfig(self.text_id, text=text)
        
    def change_color(self, bg_color):
        self.bg_color = bg_color
        self.original_bg = bg_color
        self.active_color = self.lighten_color(bg_color, 1.15)
        if not self.is_disabled:
            self.itemconfig(self.rect_id, fill=bg_color)
        
    def get_text(self):
        return self.itemcget(self.text_id, 'text')

class HoverIcon(tk.Canvas):
    def __init__(self, parent, x, y, w, h, image_name, text, command, bg_color=BG_COLOR):
        self.zoom_factor = 1.1
        self.cw = int(w * self.zoom_factor) + 10
        self.ch = int(h * self.zoom_factor) + 40
        shift_x = (self.cw - w) // 2
        shift_y = (self.ch - 40 - h) // 2
        super().__init__(parent, width=self.cw, height=self.ch, bg=bg_color, highlightthickness=0, cursor="hand2")
        self.place(x=x - shift_x, y=y - shift_y)
        self.command = command
        self.text = text
        self.w, self.h = w, h
        self.img_path = os.path.join(ICONS_DIR, image_name)
        self.has_image = os.path.exists(self.img_path)
        self.normal_img = None
        self.hover_img = None
        self.image_id = None
        self.cx = self.cw // 2
        self.cy = (self.ch - 30) // 2
        self.text_id = self.create_text(self.cx, self.ch - 15, text=text, font=("Calibri", 12, "normal"), fill=FG_COLOR)
        self.load_images(w, h, int(w * self.zoom_factor), int(h * self.zoom_factor))
        self.draw_normal()
        self.bind("<Enter>", self.on_hover)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonRelease-1>", lambda e: self.command())
        self.tag_bind(self.text_id, "<ButtonRelease-1>", lambda e: self.command())
        if self.image_id: self.tag_bind(self.image_id, "<ButtonRelease-1>", lambda e: self.command())

    def load_images(self, nw, nh, hw, hh):
        if self.has_image:
            try:
                img = Image.open(self.img_path).convert("RGBA")
                self.normal_img = ImageTk.PhotoImage(img.resize((nw, nh), Image.Resampling.LANCZOS))
                self.hover_img = ImageTk.PhotoImage(img.resize((hw, hh), Image.Resampling.LANCZOS))
            except: self.has_image = False

    def draw_normal(self):
        if self.image_id: self.delete(self.image_id)
        if self.has_image:
            self.image_id = self.create_image(self.cx, self.cy, image=self.normal_img)
        else:
            self.image_id = self.create_rectangle(10, 10, self.cw-10, self.ch-40, fill="#cccccc", outline="#999999", width=2)
        self.itemconfig(self.text_id, font=("Calibri", 12, "normal"), fill=FG_COLOR)

    def draw_hover(self):
        if self.image_id: self.delete(self.image_id)
        if self.has_image:
            self.image_id = self.create_image(self.cx, self.cy, image=self.hover_img)
        else:
            self.image_id = self.create_rectangle(5, 5, self.cw-5, self.ch-35, fill="#aaaaaa", outline="#666666", width=2)
        self.itemconfig(self.text_id, font=("Calibri", 13, "bold"), fill=ACCENT_COLOR)

    def on_hover(self, event=None):
        self.draw_hover()

    def on_leave(self, event=None):
        self.draw_normal()

def configure_treeview_style():
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Custom.Treeview", 
                    background=ENTRY_BG,
                    foreground=FG_COLOR,
                    fieldbackground=ENTRY_BG,
                    rowheight=28,
                    font=('Calibri', 12))
    style.configure("Custom.Treeview.Heading", 
                    background=BTN_BG, 
                    foreground=FG_COLOR, 
                    font=('Calibri', 13, 'bold'))
    style.map('Custom.Treeview', background=[('selected', '#5c3166')])

# --- Windows ---
def show_table_window(title, columns, data):
    win = tk.Toplevel()
    win.withdraw()  # Hide window during rendering
    win.title(title)
    win.geometry("700x500")
    win.configure(bg=BG_COLOR)
    set_titlebar_color(win, BG_COLOR)
    try: win.iconbitmap(ICON_PATH)
    except: pass
    win.geometry(f"+{(win.winfo_screenwidth() - 700) // 2}+{(win.winfo_screenheight() - 500) // 2}")

    win.transient(win.master)
    win.grab_set()
    win.focus_set()

    REGULAR_FONT = ("Segoe UI", 12)
    BOLD_FONT = ("Segoe UI Black", 12) 
    HEADER_FONT = ("Segoe UI", 11, "bold")

    frame = tk.Frame(win, bg=BG_COLOR)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    header_frame = tk.Frame(frame, bg=BG_COLOR)
    header_frame.pack(fill=tk.X, pady=(0, 2))

    # Dynamically set column tabs for perfect center alignment based on available width
    num_cols = len(columns)
    if num_cols == 2: column_widths = [330, 330]
    elif num_cols == 3: column_widths = [260, 200, 200]
    else: column_widths = [200] * num_cols

    for i, col_name in enumerate(columns):
        # Match application theme instead of white/gray
        lbl = tk.Label(header_frame, text=col_name, font=HEADER_FONT, bg=BTN_BG, fg=BTN_FG, relief="flat", pady=8)
        lbl.grid(row=0, column=i, sticky="ew", padx=1)
        header_frame.grid_columnconfigure(i, weight=1)

    # Use ENTRY_BG and ENTRY_FG to seamlessly blend with the active theme
    text_area = tk.Text(frame, wrap=tk.NONE, bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                        font=REGULAR_FONT, highlightthickness=0, pady=10)
    
    v_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text_area.yview)
    h_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=text_area.xview)
    text_area.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
    v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    text_area.pack(fill=tk.BOTH, expand=True)

    text_area.tag_configure("bold", font=BOLD_FONT, foreground=ACCENT_COLOR)

    # Configure precise center-aligned tab stops for each column
    tab_stops = []
    current_x = 0
    for width in column_widths:
        center_x = current_x + (width // 2)
        tab_stops.extend((center_x, tk.CENTER))
        current_x += width
    text_area.config(tabs=tuple(tab_stops))

    for row in data:
        text_area.insert(tk.END, "\t") # Tab to reach the first center alignment
        for i, cell_text in enumerate(row):
            cell_text = str(cell_text)
            parts = cell_text.split('*')
            for j, part in enumerate(parts):
                if j % 2 == 1: 
                    text_area.insert(tk.END, part, ("bold",))
                elif part:
                    text_area.insert(tk.END, part)
            
            if i < len(row) - 1:
                text_area.insert(tk.END, "\t")
        
        text_area.insert(tk.END, "\n\n") # Double newline for cleaner visual spacing

    text_area.config(state=tk.DISABLED)
    win.deiconify()  # Show window after setup is complete

def show_help(root):
    help_text = (
        "Spelling Injector v1.1.1 – Help & Instructions\n"
        "--------------------------------------------\n\n"
        "1) Start Practice:\n"
        "   - Initiates the main spelling practice session.\n"
        "   - The app reads a word from your active queue and plays its pronunciation.\n"
        "   - Type the correct spelling and press Enter.\n"
        "   - Fallback: If the online Google TTS engine fails (due to network), the app prompts you to switch to the built-in Offline Engine.\n\n"
        "2) Add Word (Adding from Excel):\n"
        "   - You can inject new words into the learning database using an Excel (.xlsx) file.\n"
        "   - Auto-generation: By default, a 'new.xlsx' file is automatically generated in the app folder on the first run. You can easily enter your words in the first column from top to bottom.\n"
        "   - Format Rule: Your words MUST be placed in the FIRST column (Column A).\n"
        "   - Header Rule: The app ignores the first row (e.g., 'Word'), so start your actual words from Row 2 downwards.\n"
        "   - Highlight Weak Spots: If you struggle to spell specific parts of a word, simply bold those exact letters inside your input Excel cell or 'new.xlsx' file (e.g., weird). The app will detect this and visually emphasize those letters during your practice sessions to help you memorize them faster.\n"
        "   - Adding: Click 'Add Word', select your 'new.xlsx' or your formatted Excel file, and any new words not already in the system will be injected and randomized into the queue.\n\n"
        "3) Show Queue:\n"
        "   - Displays a split view of your ongoing practice cycle.\n"
        "   - Active Queue: Words currently waiting their turn to be practiced in this cycle.\n"
        "   - Removed from Current Queue: Words you spelled correctly recently but haven't yet reached your 'target memorize count'. They are temporarily removed and will return in the next cycle.\n"
        "   - The total count of words for each category is conveniently displayed in parentheses within the headers.\n\n"
        "4) Report & Memorized:\n"
        "   - Report: A complete table of all words currently in your learning phase, showing their correct hits vs. total attempts.\n"
        "   - Memorized: Shows words that successfully reached the required target count. They are permanently memorized and logged with their 'Memorized Date'.\n"
        "   - Column headers dynamically display the total number of words and the sum of attempts for quick progress tracking.\n\n"
        "5) Download Audio:\n"
        "   - Batch downloads the pronunciation files for all words currently in your database.\n"
        "   - This ensures zero delay during practice and allows you to practice fully offline later.\n"
        "   - The download strictly follows your current Settings (Accent and Speed). For example, if you want offline access to both US and UK accents, or Slow and Normal speeds, simply adjust the Settings and click 'Download Audio' again for each configuration.\n"
        "   - Note: Your TTS Mode in Settings must be set to 'Online' to use this feature.\n\n"
        "6) Settings Menu Configuration:\n"
        "   • Target correct answers to memorize:\n"
        "     - Defines how many times you must spell a word correctly before it moves to the Memorized list (Default is 5).\n"
        "   • Pronunciation Accent:\n"
        "     - Switch between American (US) or British (UK) accents. This applies to both Online and Offline TTS.\n"
        "   • TTS Mode:\n"
        "     - 'Online (Auto)': Uses Google's realistic voice. If internet fails, it falls back to Offline.\n"
        "     - 'Force Offline': Always uses Windows built-in voices (great for offline use).\n"
        "     * IMPORTANT NOTE FOR OFFLINE VOICES:\n"
        "       To use both accents in Offline mode, you MUST install their voice packages in Windows.\n"
        "       Go to Windows Settings > Time & Language > Language & region, and add both\n"
        "       'English (United States)' and 'English (United Kingdom)'. Then verify in\n"
        "       Time & Language > Speech that both voice packages are successfully installed.\n"
        "       (Online mode works automatically and does not need this step).\n"
        "   • Speech Speed:\n"
        "     - Adjust the reading speed for both Online and Offline engines.\n"
        "   • Theme Color:\n"
        "     - Change the application's background color from a wide selection. Dark/Light mode fonts adjust automatically.\n"
        "   • Backup & Restore Database:\n"
        "     - Securely creates a timestamped folder containing a backup of your progress, and restores when needed.\n"
        "   • Clear Entire Database:\n"
        "     - Warning: This completely wipes all your current words and memorized lists and regenerates fresh, empty files.\n\n"
        "License\n"
        "   Copyright © 2026 Hojjat Sabzali\n"
        "   This project is open-source and available under the MIT License.\n"
        "\n"
        "Contact:\n"
        "   - Email: sabzali.hojjat@gmail.com\n"
        "   - LinkedIn: https://www.linkedin.com/in/hojjat-sabzali\n"
        "   - GitHub: https://github.com/HojjatSabzali/\n"
    )

    win = tk.Toplevel()
    win.withdraw()  # Hide window during rendering
    win.title("Help / About")
    win.geometry("800x650")
    win.configure(bg=BG_COLOR)
    set_titlebar_color(win, BG_COLOR)
    try: win.iconbitmap(ICON_PATH)
    except: pass
    win.geometry(f"+{(win.winfo_screenwidth() - 600) // 2}+{(win.winfo_screenheight() - 650) // 2}")
    
    win.transient(win.master)  # type: ignore
    win.grab_set()
    win.focus_set()
    
    main_frame = ttk.Frame(win)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    txt = scrolledtext.ScrolledText(main_frame, wrap="word", font=("Calibri", 12), bg=ENTRY_BG, fg=ENTRY_FG, insertbackground=ENTRY_FG)
    txt.pack(fill="both", expand=True, padx=5, pady=5)
    txt.insert(tk.END, help_text)
    
    # Text interactions setup (Emails, Links)
    email_text = "sabzali.hojjat@gmail.com"
    linkedin_text = "https://www.linkedin.com/in/hojjat-sabzali"
    github_profile_text = "https://github.com/HojjatSabzali/"
    name_text = "Hojjat Sabzali"

    def open_email(event=None): webbrowser.open(f"mailto:{email_text}")
    def open_linkedin(event=None): webbrowser.open(linkedin_text)
    def open_github_profile(event=None): webbrowser.open(github_profile_text)

    txt.tag_config("email_link", foreground="#66B2FF", underline=True)
    txt.tag_config("linkedin_link", foreground="#66B2FF", underline=True)
    txt.tag_config("github_profile_link", foreground="#66B2FF", underline=True)
    
    txt.tag_config("author_bold", font=("Calibri", 12, "bold"))
    txt.tag_config("example_bold", font=("Calibri", 12, "bold"))

    for link, tag, cmd in [(email_text, "email_link", open_email), (linkedin_text, "linkedin_link", open_linkedin), (github_profile_text, "github_profile_link", open_github_profile)]:
        start = txt.search(link, "1.0", tk.END)
        if start:
            txt.tag_add(tag, start, f"{start}+{len(link)}c")
            txt.tag_bind(tag, "<ButtonRelease-1>", cmd)
            txt.tag_bind(tag, "<Enter>", lambda e: txt.config(cursor="hand2"))
            txt.tag_bind(tag, "<Leave>", lambda e: txt.config(cursor=""))

    # Bold the author's name
    start_name = txt.search(name_text, "1.0", tk.END)
    if start_name: txt.tag_add("author_bold", start_name, f"{start_name}+{len(name_text)}c")

    # Bold the 'ei' part in the word 'weird'
    start_weird = txt.search("weird", "1.0", tk.END)
    if start_weird:
        # Adding +1c and +3c will make exactly the second and third letters, ei, bold.
        txt.tag_add("example_bold", f"{start_weird}+1c", f"{start_weird}+3c")

    context_menu = tk.Menu(txt, tearoff=0, bg=BG_COLOR, fg=FG_COLOR)
    context_menu.add_command(label="Copy", command=lambda: txt.event_generate("<<Copy>>"))
    txt.bind("<Button-3>", lambda e: context_menu.tk_popup(e.x_root, e.y_root))

    def block_keys(event):
        if (event.state & 0x4) and event.keysym.lower() in ("c", "x", "v"): return
        return "break"
    txt.bind("<Key>", block_keys)
    win.deiconify()  # Show window after setup is complete

def open_settings_window(root):
    win = tk.Toplevel(root)
    win.withdraw()  # Hide window during rendering
    win.title("Settings")
    win.geometry("500x480")
    win.configure(bg=BG_COLOR)
    set_titlebar_color(win, BG_COLOR)
    try: win.iconbitmap(ICON_PATH)
    except: pass
    win.geometry(f"+{(win.winfo_screenwidth() - 500) // 2}+{(win.winfo_screenheight() - 480) // 2}")
    win.transient(root)
    win.grab_set()
    
    config = load_config()
    
    initial_target = str(config.get("target_correct", 5))
    initial_accent = config.get("accent", "us")
    initial_tts_mode = config.get("tts_mode", "auto")
    initial_tts_speed = config.get("tts_speed", "normal")

    tk.Label(win, text="Application Settings", font=("Calibri", 20, "bold"), bg=BG_COLOR, fg=ACCENT_COLOR).pack(pady=(20, 15))

    options_frame = tk.Frame(win, bg=BG_COLOR)
    options_frame.pack(fill=tk.X, padx=30, pady=5)
    options_frame.columnconfigure(1, weight=1)

    tk.Label(options_frame, text="Correct answers to memorize:", font=("Calibri", 12), bg=BG_COLOR, fg=FG_COLOR).grid(row=0, column=0, sticky="w", pady=8)
    spinbox = tk.Spinbox(options_frame, from_=1, to=100, width=5, font=("Calibri", 12), bg=ENTRY_BG, fg=ENTRY_FG, buttonbackground=BTN_BG)
    spinbox.delete(0, tk.END)
    spinbox.insert(0, config["target_correct"])
    spinbox.grid(row=0, column=1, sticky="w", padx=10)

    tk.Label(options_frame, text="Pronunciation Accent:", font=("Calibri", 12), bg=BG_COLOR, fg=FG_COLOR).grid(row=1, column=0, sticky="w", pady=8)
    acc_var = tk.StringVar(value=config.get("accent", "us"))
    acc_frame = tk.Frame(options_frame, bg=BG_COLOR)
    acc_frame.grid(row=1, column=1, sticky="w", padx=5)
    tk.Radiobutton(acc_frame, text="US", variable=acc_var, value="us", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR).pack(side=tk.LEFT)
    tk.Radiobutton(acc_frame, text="UK", variable=acc_var, value="uk", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR).pack(side=tk.LEFT)

    tk.Label(options_frame, text="TTS Mode:", font=("Calibri", 12), bg=BG_COLOR, fg=FG_COLOR).grid(row=2, column=0, sticky="w", pady=8)
    tts_var = tk.StringVar(value=config.get("tts_mode", "auto"))
    tts_frame = tk.Frame(options_frame, bg=BG_COLOR)
    tts_frame.grid(row=2, column=1, sticky="w", padx=5)

    def on_tts_mode_change(*args):
        if tts_var.get() == "auto":
            fast_radio.config(state=tk.DISABLED)
            if spd_var.get() == "fast":
                spd_var.set("normal")
        else:
            fast_radio.config(state=tk.NORMAL)
            
    tts_var.trace_add("write", on_tts_mode_change)

    tk.Radiobutton(tts_frame, text="Online", variable=tts_var, value="auto", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR).pack(side=tk.LEFT)
    tk.Radiobutton(tts_frame, text="Offline", variable=tts_var, value="offline", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR).pack(side=tk.LEFT)

    tk.Label(options_frame, text="Speech Speed:", font=("Calibri", 12), bg=BG_COLOR, fg=FG_COLOR).grid(row=3, column=0, sticky="w", pady=8)
    spd_var = tk.StringVar(value=config.get("tts_speed", "normal"))
    spd_frame = tk.Frame(options_frame, bg=BG_COLOR)
    spd_frame.grid(row=3, column=1, sticky="w", padx=5)
    tk.Radiobutton(spd_frame, text="Slow", variable=spd_var, value="slow", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR).pack(side=tk.LEFT)
    tk.Radiobutton(spd_frame, text="Normal", variable=spd_var, value="normal", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR).pack(side=tk.LEFT)
    fast_radio = tk.Radiobutton(spd_frame, text="Fast", variable=spd_var, value="fast", bg=BG_COLOR, fg=FG_COLOR, selectcolor=ENTRY_BG, activebackground=BG_COLOR, activeforeground=FG_COLOR)
    fast_radio.pack(side=tk.LEFT)
    
    # Initialize the correct button states upon opening
    on_tts_mode_change()

    def open_color_picker():
        if hasattr(win, 'color_picker_is_open') and win.color_picker_is_open:
            return
        win.color_picker_is_open = True

        picker = tk.Toplevel(win)
        picker.withdraw()  # Hide window during rendering
        picker.title("Choose Theme Color")
        picker.geometry("400x250")
        picker.configure(bg=BG_COLOR)
        set_titlebar_color(picker, BG_COLOR)
        try: picker.iconbitmap(ICON_PATH)
        except: pass
        picker.geometry(f"+{(picker.winfo_screenwidth() - 400) // 2}+{(picker.winfo_screenheight() - 250) // 2}")
        picker.transient(win)
        picker.grab_set()
        
        def on_picker_close():
            win.color_picker_is_open = False
            picker.destroy()
            
        picker.protocol("WM_DELETE_WINDOW", on_picker_close)

        tk.Label(picker, text="Select Theme Color:", font=("Calibri", 14, "bold"), bg=BG_COLOR, fg=FG_COLOR).pack(pady=10)
        grid_frame = tk.Frame(picker, bg=BG_COLOR)
        grid_frame.pack(pady=5)

        def select_color(c):
            config["theme_color"] = c
            save_config(config)
            on_picker_close()
            messagebox.showinfo("Restart Required", "Theme color saved successfully.\nPlease restart the application to apply the changes.", parent=win)

        for r, row in enumerate(THEME_COLORS):
            for c, color in enumerate(row):
                btn = tk.Canvas(grid_frame, width=28, height=28, bg=color, highlightthickness=1, highlightbackground=FG_COLOR, cursor="hand2")
                btn.grid(row=r, column=c, padx=3, pady=3)
                btn.bind("<ButtonRelease-1>", lambda e, col=color: select_color(col))
        
        picker.deiconify()  # Show window after setup is complete

    tk.Label(options_frame, text="Theme Color:", font=("Calibri", 12), bg=BG_COLOR, fg=FG_COLOR).grid(row=4, column=0, sticky="w", pady=8)
    color_btn_frame = tk.Frame(options_frame, bg=BG_COLOR)
    color_btn_frame.grid(row=4, column=1, sticky="w", padx=5)
    RoundedButton(color_btn_frame, text="Choose Color", command=open_color_picker, bg_color=BTN_BG, fg_color=BTN_FG, width=120, height=30).pack(side=tk.LEFT)

    def save_settings():
        try:
            val = int(spinbox.get())
            if val > 0:
                config["target_correct"] = val
                config["accent"] = acc_var.get()
                config["tts_mode"] = tts_var.get()
                config["tts_speed"] = spd_var.get()
                save_config(config)
                return True
            return False
        except: 
            messagebox.showerror("Error", "Invalid number", parent=win)
            return False

    def on_save_click():
        if save_settings():
            messagebox.showinfo("Saved", "Settings saved successfully!", parent=win)
            win.destroy()

    RoundedButton(win, text="Save Settings", command=on_save_click, bg_color="#4CAF50", fg_color="white", width=250, height=35, font=("Calibri", 12, "bold")).pack(pady=(15, 10))
    
    def on_settings_close():
        current_target = spinbox.get()
        current_accent = acc_var.get()
        current_tts_mode = tts_var.get()
        current_tts_speed = spd_var.get()
        
        if (current_target != initial_target or 
            current_accent != initial_accent or 
            current_tts_mode != initial_tts_mode or 
            current_tts_speed != initial_tts_speed):
            
            response = messagebox.askyesnocancel("Unsaved Changes", 
                                                 "You have unsaved changes.\nDo you want to save them before closing?", 
                                                 parent=win)
            if response is True: # Yes (Save & Close)
                if save_settings():
                    win.destroy()
            elif response is False: # No (Ignore & Close)
                win.destroy()
            # If None (Cancel), just do nothing and leave window open
        else:
            win.destroy()

    win.protocol("WM_DELETE_WINDOW", on_settings_close)
    # -----------------------------------------------------
    
    tk.Frame(win, height=1, bg=BTN_BG).pack(fill=tk.X, padx=40, pady=5)

    btn_frame = tk.Frame(win, bg=BG_COLOR)
    btn_frame.pack(fill=tk.X, padx=30, pady=5)

    def reshuffle_queue():
        data = load_data_from_file(WORDS_PATH)
        if data['queue']:
            import random
            random.shuffle(data['queue'])
            save_data_to_file(data, WORDS_PATH)
            messagebox.showinfo("Success", "Current queue randomly reshuffled!", parent=win)
    
    def clear_database():
        if messagebox.askyesno("Warning", "Are you sure you want to delete ALL words and progress?", parent=win):
            if messagebox.askyesno("Backup", "Would you like to Backup first?\nClick Yes to Backup now, No to skip.", parent=win):
                backup_data(parent_window=win)
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'words'
                ws.append(["Word", "Correct", "Total"])
                wb.create_sheet('queue')
                wb.save(WORDS_PATH)
                wb.close()
                
                wb_comp = openpyxl.Workbook()
                ws_comp = wb_comp.active
                ws_comp.title = 'memorized'
                ws_comp.append(["Word", "Total Shown", "Memorized Date"])
                wb_comp.save(MEMORIZED_PATH)
                wb_comp.close()

                messagebox.showinfo("Reset Complete", "All data wiped and files regenerated.", parent=win)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to reset: {e}\n(Make sure files are not open)", parent=win)

    f1 = tk.Frame(btn_frame, bg=BG_COLOR)
    f1.grid(row=0, column=0, padx=5, pady=5)
    RoundedButton(f1, text="Reshuffle Queue", command=reshuffle_queue, bg_color="#2196F3", fg_color="white", width=180, height=35).pack()

    f2 = tk.Frame(btn_frame, bg=BG_COLOR)
    f2.grid(row=0, column=1, padx=5, pady=5)
    RoundedButton(f2, text="Backup Database", command=backup_data, bg_color="#9C27B0", fg_color="white", width=180, height=35).pack()

    f3 = tk.Frame(btn_frame, bg=BG_COLOR)
    f3.grid(row=1, column=0, padx=5, pady=5)
    RoundedButton(f3, text="Restore Database", command=restore_data, bg_color="#FF9800", fg_color="white", width=180, height=35).pack()

    f4 = tk.Frame(btn_frame, bg=BG_COLOR)
    f4.grid(row=1, column=1, padx=5, pady=5)
    RoundedButton(f4, text="Clear Database", command=clear_database, bg_color="#F44336", fg_color="white", width=180, height=35).pack()

    btn_frame.grid_columnconfigure(0, weight=1)
    btn_frame.grid_columnconfigure(1, weight=1)
    
    win.deiconify()  # Show window after setup is complete

# --- Practice Session ---
def create_practice_window(data):
    import time
    session_offline_fallback = False
    
    config = load_config()
    target_correct = config["target_correct"]
    words_dict = data['words']
    queue = data['queue']

    if not queue:
        candidates = [w for w, d in words_dict.items() if d['correct_count'] < target_correct]
        if not candidates:
            messagebox.showinfo("All Done!", "Amazing! All words have been memorized!")
            create_initial_window()
            return
        random.shuffle(candidates)
        queue.extend(candidates)
        save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
        messagebox.showinfo("New Cycle", "New random queue created for remaining words!")

    # Variables for current word state
    word = ""
    original_word = ""
    word_revealed = False
    is_current_word_failed = False
    last_action_time = 0

    root = tk.Tk()
    root.withdraw()  # Hide main window during rendering
    root.title("Practice Spelling")
    root.geometry("620x320")
    root.configure(bg=BG_COLOR)
    set_titlebar_color(root, BG_COLOR)
    try: root.iconbitmap(ICON_PATH)
    except: pass
    root.geometry(f"+{(root.winfo_screenwidth() - 620) // 2}+{(root.winfo_screenheight() - 320) // 2}")

    tk.Label(root, text="Spell the word you hear:", font=("Calibri", 18, "bold"), bg=BG_COLOR, fg=FG_COLOR).pack(pady=(15, 10))
    entry = tk.Entry(root, font=("Calibri", 18), width=25, justify="center", bg=ENTRY_BG, fg=ENTRY_FG, insertbackground=ENTRY_FG)
    entry.pack(pady=10)
    
    status_label = tk.Label(root, text="", font=("Calibri", 12), bg=BG_COLOR, fg="#66B2FF", width=40)
    status_label.pack(pady=5)
    
    word_display = tk.Text(root, font=("Segoe UI", 18), bg=BG_COLOR, height=1, width=35, bd=0, highlightthickness=0, state='disabled', pady=10)
    word_display.tag_configure("center", justify='center')
    
    word_display.tag_configure("normal_text", font=("Segoe UI", 18, "normal"), foreground=FG_COLOR)
    word_display.tag_configure("bold_text", font=("Segoe UI", 18, "bold"), foreground=ACCENT_COLOR)
    word_display.pack(pady=5)

    def render_word_in_display(text_word):
        word_display.config(state='normal')
        word_display.delete("1.0", tk.END)
    
        parts = text_word.split('*')
    
        for i, part in enumerate(parts):
            if i % 2 == 1:
                word_display.insert(tk.END, part, ("center", "bold_text"))
            else:
                word_display.insert(tk.END, part, ("center", "normal_text"))
            
        word_display.config(state='disabled')

    btn_replay = None
    btn_next = None
    btn_back = None

    def update_status(text, color):
        status_label.config(text=text, fg=color)
        root.update_idletasks()

    def set_buttons_state(state):
        if not btn_replay or not btn_next or not btn_back: return
        if state == "disable":
            btn_replay.disable()
            btn_next.disable()
            btn_back.disable()
        else:
            btn_replay.enable()
            btn_next.enable()
            btn_back.enable()

    def prompt_fallback():
        nonlocal session_offline_fallback
        res = messagebox.askyesno("Connection Error", "Could not connect to online TTS (Google).\nDo you want to switch to the Offline Engine for this session?")
        if res:
            session_offline_fallback = True
            btn_replay.set_text("Replay Sound")
            btn_replay.change_color("#4CAF50")
            play_word_sound()
        else:
            update_status("Sound failed. Check internet or click retry.", "#F44336")

    def handle_replay_action():
        nonlocal session_offline_fallback
        if btn_replay.is_disabled: return
        if btn_replay.get_text() == "Retry Online TTS":
            session_offline_fallback = False
        play_word_sound()

    def play_word_sound():
        set_buttons_state("disable")
        update_status("Connecting to TTS server...", "#66B2FF")
        
        def task():
            clean_tts_word = original_word.replace('*', '')
            success = play_sound(clean_tts_word, config, force_offline=session_offline_fallback)
            root.after(0, lambda: on_sound_finished(success))
            
        threading.Thread(target=task, daemon=True).start()

    def on_sound_finished(success):
        set_buttons_state("enable")
        if success:
            mode = "Offline" if (session_offline_fallback or config['tts_mode'] == 'offline') else "Online"
            update_status(f"Sound played ({mode})", "#4CAF50")
            btn_replay.set_text("Replay Sound")
            btn_replay.change_color("#4CAF50")
        else:
            if not session_offline_fallback and config['tts_mode'] != 'offline':
                btn_replay.set_text("Retry Online TTS")
                btn_replay.change_color("#FF9800") 
                prompt_fallback()
            else:
                update_status("Sound completely failed", "#F44336")
                btn_replay.set_text("Replay Sound")
                btn_replay.change_color("#4CAF50")
        
        root.after(100, lambda: entry.focus_set() if entry['state'] != 'disabled' else None)

    def load_next_word_cycle():
        nonlocal word, original_word, word_revealed, is_current_word_failed
        if not queue:
            candidates = [w for w, d in words_dict.items() if d['correct_count'] < target_correct]
            if not candidates:
                root.destroy()
                messagebox.showinfo("All Done!", "Amazing! All words have been memorized!")
                create_initial_window()
                return
            random.shuffle(candidates)
            queue.extend(candidates)
            save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
            messagebox.showinfo("New Cycle", "New random queue created for remaining words!")
        
        word = queue.pop(0)
        original_word = words_dict[word]['original_word']
        word_revealed = False
        is_current_word_failed = False
        
        entry.config(state='normal')
        entry.delete(0, tk.END)
        
        word_display.config(state='normal')
        word_display.delete("1.0", tk.END)
        word_display.config(state='disabled')
        
        btn_next.set_text("Show Spelling")
        btn_replay.set_text("Replay Sound")
        btn_replay.change_color("#4CAF50")
        
        root.unbind('<Return>')
        entry.bind('<Return>', submit)
        play_word_sound()

    def action_next_btn(event=None):
        nonlocal word_revealed, is_current_word_failed, last_action_time
        current_time = time.time()
        
        if btn_next.is_disabled or (current_time - last_action_time < 0.3): 
            return
        last_action_time = current_time

        if not word_revealed:
            render_word_in_display(original_word)
            word_revealed = True
            btn_next.set_text("Next Word")
            update_status("Spelling revealed.", "#D4AF37")
            
            if not is_current_word_failed and word in words_dict:
                words_dict[word]['total_count'] += 1
                queue.append(word)
                save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
            
            entry.delete(0, tk.END)
            entry.config(state='disabled')
            root.unbind('<Return>')
            root.bind('<Return>', action_next_btn)
        else:
            entry.config(state='normal')
            entry.delete(0, tk.END)
            root.unbind('<Return>')
            entry.bind('<Return>', submit)
            load_next_word_cycle()

    def submit(event=None):
        nonlocal word_revealed, is_current_word_failed, last_action_time
        if btn_next.is_disabled: return
        
        typed = entry.get().strip().lower()
        if not typed: return
        entry.unbind('<Return>')
        
        correct = (typed == word)
        words_dict[word]['total_count'] += 1
        
        if correct:
            words_dict[word]['correct_count'] += 1
            play_success_sound()
            if words_dict[word]['correct_count'] >= target_correct:
                add_to_memorized(original_word, words_dict[word]['total_count'])
                del words_dict[word]
                update_status("Memorized! Moved to memorized.", "#D4AF37")
            else:
                update_status("Correct spelling!", "#4CAF50")
            
            save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
            render_word_in_display(original_word)
            word_revealed = True
            btn_next.set_text("Next Word")
            
            entry.config(state='disabled')
            last_action_time = time.time()
            root.bind('<Return>', action_next_btn)
        else:
            is_current_word_failed = True
            play_error_sound()
            update_status("Wrong spelling", "#F44336")
            
            if word not in queue: queue.append(word)
            save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
            
            btn_next.set_text("Show Spelling")
            entry.delete(0, tk.END)
            entry.bind('<Return>', submit)

    def back_to_menu():
        if btn_back.is_disabled: return
        if not word_revealed and word in words_dict:
            queue.insert(0, word)
            save_data_to_file({'words': words_dict, 'queue': queue}, WORDS_PATH)
        root.destroy()
        create_initial_window()

    button_frame = tk.Frame(root, bg=BG_COLOR)
    button_frame.pack(pady=15)
    
    btn1_f = tk.Frame(button_frame, bg=BG_COLOR)
    btn1_f.pack(side=tk.LEFT, padx=5)
    btn_replay = RoundedButton(btn1_f, text="Replay Sound", command=handle_replay_action, bg_color="#4CAF50", fg_color="white", width=145, height=40)
    btn_replay.pack()
    
    btn2_f = tk.Frame(button_frame, bg=BG_COLOR)
    btn2_f.pack(side=tk.LEFT, padx=5)
    btn_next = RoundedButton(btn2_f, text="Show Spelling", command=action_next_btn, bg_color="#2196F3", fg_color="white", width=145, height=40)
    btn_next.pack()

    btn3_f = tk.Frame(button_frame, bg=BG_COLOR)
    btn3_f.pack(side=tk.LEFT, padx=5)
    btn_back = RoundedButton(btn3_f, text="Back to Menu", command=back_to_menu, bg_color="#FF9800", fg_color="white", width=145, height=40)
    btn_back.pack()

    root.after(200, entry.focus_set)
    entry.bind('<Return>', submit)
    root.protocol("WM_DELETE_WINDOW", back_to_menu)
    root.after(200, load_next_word_cycle)
    root.deiconify()  # Show main window after setup is complete
    root.mainloop()

def download_all_pronunciations(root):
    config = load_config()
    if config.get("tts_mode", "auto") == "offline":
        messagebox.showinfo("Offline Mode Active", 
                            "Your TTS mode is currently set to 'Offline', so downloading is not required.\n\n"
                            "If you wish to download audio files, please go to Settings, switch to 'Online' mode, select your desired accent and speed, and then click Download Audio.",
                            parent=root)
        return

    data = load_data_from_file(WORDS_PATH)
    words_dict = data['words']
    if not words_dict:
        messagebox.showinfo("Empty", "No words found in database to download.", parent=root)
        return

    accent = config.get("accent", "us").strip().lower()
    tld_val = 'us' if accent == 'us' else 'co.uk'
    speed_setting = config.get("tts_speed", "normal")
    is_slow = (speed_setting == "slow")
    speed_suffix = "slow" if is_slow else "normal"

    # Check which words have not been downloaded yet (considering the accent and speed)
    words_to_download = []
    for word in words_dict.keys():
        clean_word = word.replace('*', '').strip().lower()
        file_path = os.path.join(PRONUNCIATIONS_DIR, f"{clean_word}_{accent}_{speed_suffix}.mp3")
        if not os.path.exists(file_path):
            words_to_download.append(clean_word)

    if not words_to_download:
        messagebox.showinfo("Done!", f"All pronunciations ({accent.upper()} - {speed_suffix.title()}) are already downloaded!", parent=root)
        return

    # Create the download window
    dl_win = tk.Toplevel(root)
    dl_win.withdraw()  # Hide window during rendering
    dl_win.title("Downloading Audio")
    dl_win.geometry("450x250")
    dl_win.configure(bg=BG_COLOR)
    set_titlebar_color(dl_win, BG_COLOR)
    try: dl_win.iconbitmap(ICON_PATH)
    except: pass
    dl_win.geometry(f"+{(dl_win.winfo_screenwidth() - 450) // 2}+{(dl_win.winfo_screenheight() - 250) // 2}")
    dl_win.transient(root)
    dl_win.grab_set()

    tk.Label(dl_win, text="Downloading Missing Pronunciations", font=("Calibri", 16, "bold"), bg=BG_COLOR, fg=FG_COLOR).pack(pady=(15, 5))
    tk.Label(dl_win, text=f"Target: {accent.upper()} Accent | Speed: {speed_suffix.title()}", font=("Calibri", 12), bg=BG_COLOR, fg="#66B2FF").pack(pady=(0, 10))
    
    progress_lbl = tk.Label(dl_win, text=f"0 / {len(words_to_download)}", font=("Calibri", 14, "bold"), bg=BG_COLOR, fg=ACCENT_COLOR)
    progress_lbl.pack(pady=5)
    
    word_lbl = tk.Label(dl_win, text="Preparing...", font=("Calibri", 12), bg=BG_COLOR, fg=FG_COLOR)
    word_lbl.pack(pady=5)
    
    progress_bar = ttk.Progressbar(dl_win, orient=tk.HORIZONTAL, length=350, mode='determinate', maximum=len(words_to_download))
    progress_bar.pack(pady=10)
    dl_win.deiconify()  # Show window after setup is complete

    def download_task():
        downloaded_count = 0
        for clean_word in words_to_download:
            if not dl_win.winfo_exists(): break  # Stop if the user closes the window
            
            word_lbl.config(text=f"Fetching: {clean_word}")
            dl_win.update()
            
            file_path = os.path.join(PRONUNCIATIONS_DIR, f"{clean_word}_{accent}_{speed_suffix}.mp3")
            try:
                tts = gTTS(text=clean_word, lang='en', tld=tld_val, slow=is_slow)
                tts.save(file_path)
            except Exception as e:
                print(f"Failed to download '{clean_word}': {e}")
            
            downloaded_count += 1
            progress_lbl.config(text=f"{downloaded_count} / {len(words_to_download)}")
            progress_bar['value'] = downloaded_count
            dl_win.update()
            
            # Short delay to prevent being blocked by Google
            import time
            time.sleep(0.4) 

        if dl_win.winfo_exists():
            word_lbl.config(text="All downloads completed successfully!")
            dl_win.update()
            messagebox.showinfo("Complete", "Download process finished successfully!", parent=dl_win)
            dl_win.destroy()

    threading.Thread(target=download_task, daemon=True).start()

# --- Main Menu UI ---
def create_initial_window():
    root = tk.Tk()
    root.withdraw()  # Hide main window during rendering
    root.title("Spelling Injector")
    root.geometry("800x550")
    root.configure(bg=BG_COLOR)
    set_titlebar_color(root, BG_COLOR)
    try: root.iconbitmap(ICON_PATH)
    except: pass
    root.geometry(f"+{(root.winfo_screenwidth() - 800) // 2}+{(root.winfo_screenheight() - 550) // 2}")

    def on_closing():
        root.destroy()
        try:
            import pygame
            pygame.quit()
        except:
            pass
        import os
        os._exit(0) 
        
    root.protocol("WM_DELETE_WINDOW", on_closing)

    canvas = tk.Canvas(root, width=800, height=550, bg=BG_COLOR, highlightthickness=0)
    canvas.pack(fill=tk.BOTH, expand=True)

    HoverIcon(canvas, 680, 30, 80, 40, "settings.png", "Settings", lambda: open_settings_window(root), bg_color=BG_COLOR)
    HoverIcon(canvas, 580, 30, 80, 40, "help.png", "Help", lambda: show_help(root), bg_color=BG_COLOR)

    def start_practice():
        data = load_data_from_file(WORDS_PATH)
        if not data['words']:
            messagebox.showinfo("No Words", "Please add words first!")
            return
        root.destroy()
        create_practice_window(data)

    def show_queue():
        data = load_data_from_file(WORDS_PATH)
        words_dict = data['words']
        q_list = data['queue']
        active_set = set(q_list)
        cfg = load_config()
        t_corr = cfg['target_correct']
        
        temp_removed = [d['original_word'] for k, d in words_dict.items() if k not in active_set and d['correct_count'] < t_corr]
        
        max_len = max(len(q_list), len(temp_removed))
        table_data = []
        
        active_count = len(q_list)
        removed_count = len(temp_removed)
        
        for i in range(max_len):
            col1 = words_dict[q_list[i]]['original_word'] if i < len(q_list) else ""
            col2 = temp_removed[i] if i < len(temp_removed) else ""
            table_data.append((col1, col2))

        cols = [f"Active Queue ({active_count})", f"Removed from Current Queue ({removed_count})"]
        show_table_window("Current Practice Queue", cols, table_data)

    def view_report():
        data = load_data_from_file(WORDS_PATH)
        words_dict = data['words']
        if not words_dict:
            messagebox.showinfo("Report", "No words in learning process.")
            return
        table_data = [(d['original_word'], d['correct_count'], d['total_count']) for d in words_dict.values()]
        
        total_words = len(words_dict)
        sum_correct = sum(d['correct_count'] for d in words_dict.values())
        sum_total = sum(d['total_count'] for d in words_dict.values())
        
        cols = [f"Word ({total_words})", f"Correct Hits ({sum_correct})", f"Total Attempts ({sum_total})"]
        show_table_window("Learning Report", cols, table_data)

    def view_memorized():
        try:
            comp_wb = openpyxl.load_workbook(MEMORIZED_PATH, data_only=True)
            sheet = comp_wb.active
            table_data = []
            sum_attempts = 0
            
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                if row[0]: 
                    table_data.append((row[0], row[1], row[2]))
                    try: sum_attempts += int(row[1])
                    except: pass
            comp_wb.close()
            
            if not table_data:
                messagebox.showinfo("Memorized", "No memorized words yet.")
                return
            
            total_words = len(table_data)
            cols = [f"Word ({total_words})", f"Total Attempts ({sum_attempts})", "Memorized Date"]
            show_table_window("Memorized Words", cols, table_data)
        except FileNotFoundError:
            messagebox.showinfo("Memorized", "No memorized words yet.")

    # Main primary action
    HoverIcon(canvas, 70, 200, 220, 150, "start.png", "Start Practice", start_practice, bg_color=BG_COLOR)
    
    # Center queue block
    HoverIcon(canvas, 340, 150, 120, 250, "queue.png", "Show Queue", show_queue, bg_color=BG_COLOR)
    
    # Column 1: Add Word & Report (Centered relative to each other)
    HoverIcon(canvas, 490, 130, 140, 140, "add.png", "Add Word", load_words_from_xlsx, bg_color=BG_COLOR)
    HoverIcon(canvas, 505, 320, 110, 110, "report.png", "Report", view_report, bg_color=BG_COLOR)
    
    # Column 2: Download Audio & Memorized (Shifted right to prevent overlap)
    HoverIcon(canvas, 660, 145, 110, 110, "download_audio.png", "Download Audio", lambda: download_all_pronunciations(root), bg_color=BG_COLOR)
    HoverIcon(canvas, 660, 320, 110, 110, "memorized.png", "Memorized", view_memorized, bg_color=BG_COLOR)
    root.deiconify()  # Show main window after setup is complete
    root.mainloop()

def main():
    # Prevent multiple instances
    mutex = ctypes.windll.kernel32.CreateMutexW(None, False, "SpellingInjector_Mutex")
    if ctypes.windll.kernel32.GetLastError() == 183: # ERROR_ALREADY_EXISTS
        hwnd = ctypes.windll.user32.FindWindowW(None, "Spelling Injector")
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 9) # SW_RESTORE
            ctypes.windll.user32.SetForegroundWindow(hwnd)
        sys.exit(0)

    if not os.path.exists(ICONS_DIR): os.makedirs(ICONS_DIR)
    pygame.mixer.init(frequency=44100, size=-16, channels=2, buffer=512)
    create_empty_db_files()
    create_initial_window()
    
    pygame.quit()
    os._exit(0)

if __name__ == "__main__":
    main()
